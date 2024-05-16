import pandas as pd
import numpy as np
import datetime
import matplotlib.pyplot as plt
from qiskit_ibm_runtime import SamplerV2 as Sampler
from qiskit_finance.circuit.library.probability_distributions import NormalDistribution
from qiskit import QuantumCircuit
from qiskit.primitives import Sampler
from qiskit.circuit.library import QFT
from qiskit_finance.data_providers import BaseDataProvider
from qiskit_finance.circuit.library import GaussianConditionalIndependenceModel
from qiskit_ibm_runtime import QiskitRuntimeService, Session

from qiskit.quantum_info import Statevector


class StockDataProcessor(BaseDataProvider):
    """
    StockDataProcessor is a child class of parent class BaseDataProvider from Qiskit Finance. 
    Storing data in this form will be beneficial as it allows usage of further Qiskit Finance functions. 

    """
    def __init__(self, file_path, start, end):
        """
        Initialize with the path to the Excel file and date range.
        """
        self._file_path = file_path
        self._start = pd.to_datetime(start)
        self._end = pd.to_datetime(end)
        self._tickers = []
        self._data = pd.DataFrame()
    def load_data(self) -> pd.DataFrame:
        """
        Loads data from an Excel file into a DataFrame and sets the tickers.
        """
        try:
            # Read data from the provided Excel file
            df = pd.read_excel(self._file_path, index_col=0)
            df.index = pd.to_datetime(df.index)
            df.sort_values("Date")
            df.reset_index(drop=True)

            # Filter the DataFrame to only include the specified date range
            df = df[(df.index >= self._start) & (df.index <= self._end)]
            # Set the tickers to the column headers
            self._tickers = df.columns.tolist()
            return df
        except Exception as e:
            raise IOError(f"Error loading data from {self._file_path}: {e}")

    @staticmethod
    def calculate_log_returns(prices: pd.Series) -> pd.Series:
        """
        Calculates the log returns for a series of prices.

        Parameters:
        prices (pd.Series): A series of percentage returns or prices.

        Returns:
        pd.Series: A series containing log returns.
        """
        return np.log(1 + prices).dropna()

    def run(self) -> None:
        """
        Processes each ticker to calculate log returns from the Excel file.
        """
        # Load data from the Excel file
        df = self.load_data()
        
        # Calculate log returns for all tickers
        log_returns = df.apply(self.calculate_log_returns, axis=0)
        
        # Drop rows with NaN values (e.g., first row due to shift)
        self._data = log_returns.dropna()

    def get_period_return_mean_vector(self) -> np.ndarray:
        """
        Computes the mean vector of period returns.

        Returns:
        np.ndarray: Mean vector of period returns.
        """
        if self._data.empty:
            raise ValueError("No data available. Ensure `run` has been executed successfully.")

        return self._data.mean(axis=0).to_numpy()

    def get_period_return_covariance_matrix(self) -> np.ndarray:
        """
        Computes the covariance matrix of period returns.

        Returns:
        np.ndarray: Covariance matrix of period returns.
        """
        if self._data.empty:
            raise ValueError("No data available. Ensure `run` has been executed successfully.")

        return self._data.cov().to_numpy()

#        return np.log(1 + prices).dropna()

# Example usage of class
data = StockDataProcessor(
    start=datetime.datetime(2004, 4, 30),
    end=datetime.datetime(2024, 3, 31),
    file_path="./data/historic_data.xlsx"
)


data.run()

# Plotting log returns for each ticker
for (cnt, ticker) in enumerate(data._tickers):
    plt.plot(data._data.iloc[:, cnt], label=ticker)
plt.legend(loc="upper center", bbox_to_anchor=(0.5, 1.1), ncol=3)
plt.xticks(rotation=90)
plt.savefig('StockGraph.png')


mean_vector = data.get_period_return_mean_vector() 
cov_matrix = data.get_period_return_covariance_matrix() 
precision_matrix = np.linalg.inv(cov_matrix)  
volatility = np.exp(np.sqrt(np.diag(cov_matrix))) - 1

std_devs = np.sqrt(np.diag(cov_matrix))
# Calculate the correlation matrix
correlation_matrix = cov_matrix / np.outer(std_devs, std_devs)
print(correlation_matrix)
flattened = list(set(correlation_matrix.flatten()))
correlation_matrix_adjusted = [ round(elem, 4) for elem in flattened ] # Sensitivities of default probabilities rounded to 4 decimal places
correlation_matrix_adjusted.remove(1.) #remove dummy data (correlation of 1 for same stock)
correlation_matrix_adjusted.remove(1.)
correlation_matrix_adjusted.remove(1.)

annual_expected_returns = np.array([0.1, 0.1, 0.06]) # Standard default probabilities
monthly_expected_log_returns = np.log(1 + annual_expected_returns) / 12

import seaborn as sns


simulated_log_returns = np.random.multivariate_normal(
    monthly_expected_log_returns, cov_matrix, 360
)
simulated_log_returns = pd.DataFrame(
    simulated_log_returns,
    columns=["US Equities", "International Equities", "Global Fixed Income"],
)
# Set up the matplotlib figure
plt.figure(figsize=(18, 6))

# Draw a subplot for each asset category
for i, column in enumerate(simulated_log_returns.columns, 1):
    plt.subplot(1, 3, i)  # 1 row, 3 columns, ith subplot
    sns.histplot(simulated_log_returns[column], bins=30, kde=True)
    plt.title(f'Distribution of {column}')
    plt.xlabel('Log Returns')
    plt.ylabel('Frequency')

plt.tight_layout()
plt.savefig("expectedoutput.png")

num_qubits = [4,4,4]
print(monthly_expected_log_returns)
print(cov_matrix)
std_devs = np.sqrt(np.diag(cov_matrix))

print(std_devs)
# Calculate bounds as +- 3 standard deviations around the mean
bounds = [(monthly_expected_log_returns[i] - 3*std_devs[i], monthly_expected_log_returns[i] + 3*std_devs[i]) for i in range(len(monthly_expected_log_returns))]

# Print calculated bounds
print("Calculated Bounds:")
for i, b in enumerate(bounds):
    print(f"Dimension {i+1}: {b}")
mvnd = NormalDistribution(num_qubits,monthly_expected_log_returns, cov_matrix, bounds=bounds )

qc = QuantumCircuit(sum(num_qubits))
qc.append(mvnd, range(sum(num_qubits)))
#qc.append(QFT(sum(num_qubits)), range(sum(num_qubits)))
qc.measure_all()


# Sample using the Sampler primitive
sampler = Sampler()
job = sampler.run([qc], shots=360)
result = job.result()

# Extract quasi-probabilities and convert them to binary-encoded samples
counts = result.quasi_dists[0].nearest_probability_distribution().binary_probabilities()
print(counts)
binary_samples = [k for k, v in counts.items() for _ in range(int(v * 360))]

# Decode samples back to individual asset values
def binary_to_asset_values(binary_sample, num_qubits, mu, sigma):
    asset_values = []
    start_idx = 0 # Index to keep track of qubit groups
    for i, qubits in enumerate(num_qubits):
        end_idx = start_idx + qubits # End index for current asset's qubits
        asset_bin = binary_sample[start_idx:end_idx] # Get the binary string
         # Convert binary to float in [0, 1] range and scale to asset return
        asset_value = int(asset_bin, 2) / (2**qubits - 1)
        value = mu[i] + np.sqrt(sigma[i][i]) * (2 * asset_value - 1)
        asset_values.append(value)
        start_idx = end_idx # Move to the next set of qubits
    return asset_values

# Apply the conversion function to all samples
asset_samples = np.array([binary_to_asset_values(sample, num_qubits, monthly_expected_log_returns, cov_matrix) for sample in binary_samples])


def create_new_xslx_monthly_dates(load_data, filename):
    import openpyxl
    import datetime
    import calendar
    import os

    def month_increment(start_date, num_months):
        # Calculate the new month and year
        new_month = (start_date.month + num_months - 1) % 12 + 1
        new_year = start_date.year + (start_date.month + num_months - 1) // 12
        
        # Calculate the last day of the new month
        last_day_of_month = calendar.monthrange(new_year, new_month)[1]
        
        # Ensure the new day is the last valid day of the new month if the original day doesn't exist in the new month
        new_day = min(start_date.day, last_day_of_month)
        return datetime.date(new_year, new_month, new_day)

    # Define the start date
    start_date = datetime.date(2004, 4, 30)

    # Generate monthly dates for each row in the data array
    monthly_dates = [month_increment(start_date, i) for i in range(load_data.shape[0])]

    if os.path.exists(filename):
            wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active

    # Clear the existing data in the worksheet
    ws.delete_rows(1, ws.max_row)
    # Create a new workbook and select the active worksheet
   

    # Set the column labels
    ws.append(['Date', '^GSPC', '^ACWX', '^GLAB.L'])

    # Iterate over the data and append each row to the worksheet with the monthly date
    for i, row in enumerate(load_data):
        ws.append([monthly_dates[i].strftime('%Y-%m-%d')] + row.tolist())

    # Save the workbook
    wb.save(filename)


create_new_xslx_monthly_dates(asset_samples,filename="output.xslx")

run_data = StockDataProcessor( 
    start=datetime.datetime(2004, 4, 30),
    end=datetime.datetime(2024, 3, 31),
    file_path="output2.xlsx")
run_data.run()
cov_matrix = run_data.get_period_return_covariance_matrix() 
print(cov_matrix)

std_devs = np.sqrt(np.diag(cov_matrix))
# Calculate the correlation matrix
correlation_matrix2 = cov_matrix / np.outer(std_devs, std_devs)
print(correlation_matrix2)
# Plot the sampled distribution
fig, axes = plt.subplots(1, 3, figsize=(18, 6))
for i, asset in enumerate(data._tickers):
    sns.histplot(asset_samples[:, i], bins=15, kde=False, ax=axes[i], color='blue')
    axes[i].set_xlabel(f'{asset} Returns')
    axes[i].set_ylabel('Frequency')
    axes[i].set_title(f'{asset} Returns Distribution (120 Samples)')

fig.suptitle('Sample Distribution of Multivariate Normal Distribution (120 Samples)')
plt.show()

