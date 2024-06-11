import openpyxl
import calendar
import os
import datetime
import numpy as np
from scipy.stats import norm
from scipy.linalg import cholesky
from scipy.stats import norm


#make 3 lists
# count how many times you have 000, 001, 010, 011, 100, 101, 110, 111 for the first list, which 

#def build_asset_bin_list(binary_sample, num_qubits):
    #for i, qubits in enumerate(num_qubits):


def binary_to_asset_values_timeseries(time_series, mu, sigma):
    elements1 = [sublist[0] for sublist in time_series]
    elements2 = [sublist[1] for sublist in time_series]
    elements3 = [sublist[2] for sublist in time_series]
    average1 = np.mean(elements1)
    average2 = np.mean(elements2)
    average3 = np.mean(elements3)
    stddev1 = np.std(elements1)
    stddev2 = np.std(elements2)
    stddev3 = np.std(elements3)


    for row in time_series:
        row[0] = ((row[0]- average1) / stddev1 ) * np.sqrt(sigma[0][0]) + mu[0]
        row[1] = ((row[1]- average2) / stddev2 ) * np.sqrt(sigma[1][1]) + mu[1]
        row[2] = ((row[2]- average3) / stddev3 ) * np.sqrt(sigma[2][2]) + mu[2]
        #row[0] = np.log(1+ ((row[0]- average1) / stddev1 ) * np.sqrt(sigma[0][0]) + mu[0])
        #row[1] = np.log(1+((row[1]- average2) / stddev2 ) * np.sqrt(sigma[1][1]) + mu[1])
        #row[2] = np.log(1+((row[2]- average3) / stddev3 ) * np.sqrt(sigma[2][2]) + mu[2])

    

def binary_to_asset_values_redone(dict1, dict2, dict3, num_qubits, mu, sigma):
    list1 = []
    list2 = []
    list3 = []
    #convert all dictionary keys from binary strings to ints
    dict1 = {int(str(key), 2): val for key, val in dict1.items()}
    dict2 = {int(str(key), 2): val for key, val in dict2.items()}
    dict3 = {int(str(key), 2): val for key, val in dict3.items()}
    print(dict1.keys())
    start_idx = 0 # Index to keep track of qubit groups
    for i, qubits in enumerate(num_qubits):
        start_idx += qubits # End index for current asset's qubits
        if(start_idx == 3):
            for key, val in dict1.items():
                value = ((key - np.average(list(dict1.keys()))) / np.std(list(dict1.keys()))) * np.sqrt(sigma[i][i]) + mu[i]    
                list1.append(value)
        elif(start_idx == 6):
            for key, val in dict2.items():
                value = ((key - np.average(list(dict2.keys()))) / np.std(list(dict2.keys()))) * np.sqrt(sigma[i][i]) + mu[i]                
                list2.append(value)
        else:
            for key, val in dict3.items():
                value = ((key - np.average(list(dict3.keys()))) / np.std(list(dict3.keys()))) * np.sqrt(sigma[i][i]) + mu[i]                
                list3.append(value)

    return list1,list2,list3


def binary_to_asset_values(binary_sample, num_qubits, mu, sigma):
    asset_values = []
    start_idx = 0 # Index to keep track of qubit groups
    print(binary_sample)
    for i, qubits in enumerate(num_qubits):
        end_idx = start_idx + qubits # End index for current asset's qubits
        asset_bin = binary_sample[start_idx:end_idx] # Get the binary string
        #print(binary_sample)
        #print(len(binary_sample))
        print("Asset bin: ",asset_bin)
        # Convert binary to float in [0, 1] range and scale to asset return
        asset_value = int(asset_bin, 2) #/ (2**qubits - 1) #dont need the divison

        #asset_value = ((asset_value - avg of that asset) / sigma of the entire list of that asset class) * np.sqrt(sigma[i][i]) + mu[i] 
        print("Asset value: " , asset_value)
        value = mu[i] + np.sqrt(sigma[i][i]) * (6 * asset_value - 3) #(2.3 * z_value) #(4 * asset_value - 2)# 
        print("Final value: " , value)
        print("\n\n")
        asset_values.append(value)
        start_idx = end_idx # Move to the next set of qubits
    return asset_values

def create_new_xlsx_monthly_dates(load_data, filename, secondTime = 0):

    def month_increment(start_date, num_months):
        # Calculate the new month and year
        new_month = (start_date.month + num_months - 1) % 12 + 1
        new_year = start_date.year + (start_date.month + num_months - 1) // 12
        
        # Calculate the last day of the new month
        last_day_of_month = calendar.monthrange(new_year, new_month)[1]
        
        # Ensure the new day is the last valid day of the new month if the original day doesn't exist in the new month
        new_day = min(start_date.day, last_day_of_month)
        return datetime.date(new_year, new_month, new_day)
    start_date = datetime.date(2024, 4, 30)
    monthly_dates = [month_increment(start_date, i) for i in range(load_data.shape[0])]

    if os.path.exists(filename):
        wb = openpyxl.load_workbook(filename)
    else:
        wb = openpyxl.Workbook()

    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(['Date', '^GSPC', '^ACWX', '^GLAB.L'])

    for i, row in enumerate(load_data):
        ws.append([monthly_dates[i].strftime('%Y-%m-%d')] + row.tolist())
    wb.save(filename)

